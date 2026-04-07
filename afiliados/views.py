from django.shortcuts import render, redirect
import os
from django.contrib.auth import authenticate, login
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.core.paginator import Paginator
from django.db.models import Q
from .models import Afiliado
from datetime import datetime
from django.shortcuts import render, redirect
from django.contrib import messages
import pandas as pd
from datetime import datetime
from .models import Afiliado
from django.shortcuts import get_object_or_404, render
from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Q
from .models import Afiliado, Atencion, Servicio
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
import json
# views.py
from django.shortcuts import render
from django.contrib import messages
from django.http import JsonResponse
import pandas as pd
from .models import *
from django.contrib.auth.decorators import user_passes_test
import openpyxl
import re
from django.contrib.auth.hashers import make_password
from django.db import transaction
from django.utils import timezone

@csrf_exempt
@login_required
def guardar_observacion(request, afiliado_id):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Método no permitido'}, status=405)

    try:
        afiliado = Afiliado.objects.get(id=afiliado_id)
        data = json.loads(request.body)
        campo = data.get('campo')
        valor = data.get('valor', '')

        # Validar campo permitido
        campos_permitidos = [
            'observacion_enfermeria',
            'observacion_odontologia',
            'observacion_obstetricia'
        ]
        if campo not in campos_permitidos:
            return JsonResponse({'success': False, 'message': 'Campo no válido'}, status=400)

        # Actualizar el campo dinámicamente
        setattr(afiliado, campo, valor)
        afiliado.save()

        return JsonResponse({
            'success': True,
            'message': f'Observación {campo.replace("observacion_", "")} actualizada correctamente'
        })
    except Afiliado.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Afiliado no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

@login_required
@csrf_exempt
@require_http_methods(["POST"])
def eliminar_servicio_cumplido(request, afiliado_id):
    """
    View para eliminar un servicio marcado como cumplido (eliminar la atención)
    Solo elimina atenciones del año actual
    """
    try:
        data = json.loads(request.body)
        servicio_nombre = data.get('servicio')

        if not servicio_nombre:
            return JsonResponse({
                'success': False,
                'message': 'Nombre del servicio es requerido'
            }, status=400)

        # Obtener afiliado y servicio
        try:
            afiliado = Afiliado.objects.get(id=afiliado_id)
            servicio = Servicio.objects.get(nombre=servicio_nombre)
        except (Afiliado.DoesNotExist, Servicio.DoesNotExist):
            return JsonResponse({
                'success': False,
                'message': 'Afiliado o servicio no encontrado'
            }, status=404)

        # Obtener el año actual
        año_actual = timezone.now().year

        # Buscar la atención más reciente de ESTE AÑO para este servicio y afiliado
        atencion = Atencion.objects.filter(
            afiliado=afiliado,
            servicio=servicio,
            fecha_atencion__year=año_actual  # 🔥 Solo del año actual
        ).order_by('-fecha_atencion').first()

        if not atencion:
            return JsonResponse({
                'success': False,
                'message': f'No se encontró una atención del año {año_actual} para eliminar'
            }, status=404)

        # Guardar información para el mensaje antes de eliminar
        servicio_nombre = atencion.servicio.nombre
        fecha_atencion = atencion.fecha_atencion.date()

        # Eliminar la atención
        atencion.delete()

        return JsonResponse({
            'success': True,
            'message': f'Servicio {servicio_nombre} del {fecha_atencion} eliminado correctamente'
        })

    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'Error en el formato de datos JSON'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error del servidor: {str(e)}'
        }, status=500)

@login_required
@csrf_exempt
@require_http_methods(["POST"])
def marcar_servicio_cumplido(request, afiliado_id):
    """
    View para marcar un servicio como cumplido para un afiliado
    y registrar quién lo hizo (usuario autenticado)
    """
    try:
        data = json.loads(request.body)
        servicio_nombre = data.get('servicio')
        observaciones = data.get('observaciones', '')

        if not servicio_nombre:
            return JsonResponse({
                'success': False,
                'message': 'Nombre del servicio es requerido'
            }, status=400)

        # Obtener afiliado y servicio
        try:
            afiliado = Afiliado.objects.get(id=afiliado_id)
            servicio = Servicio.objects.get(nombre=servicio_nombre)
        except (Afiliado.DoesNotExist, Servicio.DoesNotExist):
            return JsonResponse({
                'success': False,
                'message': 'Afiliado o servicio no encontrado'
            }, status=404)

        # Verificar si ya existe una atención para hoy (opcional)
        hoy = timezone.now().date()
        if Atencion.objects.filter(
            afiliado=afiliado,
            servicio=servicio,
            fecha_atencion__date=hoy
        ).exists():
            return JsonResponse({
                'success': False,
                'message': 'Ya existe una atención registrada para este servicio hoy'
            }, status=400)

        # Crear la atención y registrar quién la hizo
        atencion = Atencion.objects.create(
            afiliado=afiliado,
            servicio=servicio,
            observaciones=observaciones,
            registrado_por=request.user  # 🔥 guarda al usuario autenticado
        )

        return JsonResponse({
            'success': True,
            'message': f'Servicio {servicio.nombre} marcado como cumplido por {request.user.username}',
            'atencion_id': atencion.id
        })

    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'Error en el formato de datos JSON'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error del servidor: {str(e)}'
        }, status=500)




from django.db.models import Count
from datetime import datetime, timedelta
from django.utils import timezone
from django.db.models import Count, Q
from datetime import datetime, timedelta
from django.utils import timezone
from django.db import models
@login_required
def reportes_view(request):
    """
    Vista para reportes de cumplimiento mensual
    """
    # Obtener mes y año actual
    hoy = timezone.now().date()
    mes_actual = hoy.month
    año_actual = hoy.year

    # Calcular fechas del mes actual
    fecha_inicio_mes = hoy.replace(day=1)
    fecha_fin_mes = (fecha_inicio_mes + timedelta(days=32)).replace(day=1) - timedelta(days=1)

    # Convertir a datetime para consultas
    fecha_inicio_dt = timezone.make_aware(datetime.combine(fecha_inicio_mes, datetime.min.time()))
    fecha_fin_dt = timezone.make_aware(datetime.combine(fecha_fin_mes, datetime.max.time()))

    # 1. Total de pacientes atendidos este mes
    pacientes_atendidos_mes = Afiliado.objects.filter(
        atenciones__fecha_atencion__range=(fecha_inicio_dt, fecha_fin_dt)
    ).distinct().count()

    # 2. Obtener todos los servicios disponibles
    todos_servicios = Servicio.objects.all()

    # 3. Afiliados que ya tienen TODAS las atenciones del mes
    afiliados_completos = []
    afiliados_atendidos_mes = Afiliado.objects.filter(
        atenciones__fecha_atencion__range=(fecha_inicio_dt, fecha_fin_dt)
    ).distinct()

    for afiliado in afiliados_atendidos_mes:
        servicios_atendidos = afiliado.servicios_del_anio(año_actual)
        if servicios_atendidos.count() == todos_servicios.count():
            afiliados_completos.append(afiliado)

    total_completos = len(afiliados_completos)

    # 4. Datos para la tabla: afiliados atendidos este mes con estado de cada servicio
    afiliados_con_estado = []
    
    # Contadores para pacientes con atenciones pendientes por etapa de vida
    faltan_nino = 0
    faltan_adolescente = 0
    faltan_joven = 0
    faltan_adulto = 0
    faltan_mayor = 0

    for afiliado in afiliados_atendidos_mes:
        servicios_atendidos = set(afiliado.servicios_del_anio(año_actual))
        servicios_pendientes = set(todos_servicios) - servicios_atendidos

        # Crear lista de servicios con estado para el template
        servicios_estado = []
        for servicio in todos_servicios:
            servicios_estado.append({
                'nombre': servicio.nombre,
                'cumplido': servicio in servicios_atendidos
            })

        completo = len(servicios_pendientes) == 0

        # Clasificación de edad si le faltan atenciones (y tiene fecha de nacimiento)
        if not completo and afiliado.fecha_nacimiento:
            nac = afiliado.fecha_nacimiento
            # Calcular edad precisa considerando meses y días
            anios = hoy.year - nac.year - ((hoy.month, hoy.day) < (nac.month, nac.day))
            
            if anios < 12:
                faltan_nino += 1
            elif anios < 18:
                faltan_adolescente += 1
            elif anios < 30:
                faltan_joven += 1
            elif anios < 60:
                faltan_adulto += 1
            else:
                faltan_mayor += 1

        afiliados_con_estado.append({
            'afiliado': afiliado,
            'servicios_estado': servicios_estado,
            'total_atendidos': len(servicios_atendidos),
            'total_pendientes': len(servicios_pendientes),
            'completo': completo
        })

    context = {
        'pacientes_atendidos_mes': pacientes_atendidos_mes,
        'total_completos': total_completos,
        'afiliados_con_estado': afiliados_con_estado,
        'todos_servicios': todos_servicios,
        'mes_actual': fecha_inicio_mes.strftime("%B %Y"),
        'fecha_inicio_mes': fecha_inicio_mes,
        'fecha_fin_mes': fecha_fin_mes,
        'faltan_nino': faltan_nino,
        'faltan_adolescente': faltan_adolescente,
        'faltan_joven': faltan_joven,
        'faltan_adulto': faltan_adulto,
        'faltan_mayor': faltan_mayor,
    }

    return render(request, 'reportes.html', context)
@user_passes_test(lambda u: u.is_superuser)
def admin_panel(request):
    return render(request, "admin.html")

@login_required
def atenciones_create(request):
    if request.method == "POST":
        try:
            afiliado_id = request.POST.get("afiliado")
            servicio_id = request.POST.get("servicio")
            obs = request.POST.get("observaciones", "")

            # ejemplo guardado
            atencion = Atencion.objects.create(
                afiliado_id=afiliado_id,
                servicio_id=servicio_id,
                observaciones=obs
            )

            return JsonResponse({"success": True, "id": atencion.id})
        except Exception as e:
            return JsonResponse({"success": False, "message": str(e)})

    return JsonResponse({"success": False, "message": "Método no permitido"})


def cargar_afiliados_excel(request):
    """
    Controlador de carga masiva Multi-paso:
    1. Upload inicial -> Retorna Preview + Total
    2. Peticiones AJAX -> Procesa bloques (batches) con barra de progreso
    """
    if request.method == "POST":
        # CASO 2: PROCESAMIENTO POR LOTES (Vía AJAX)
        if request.headers.get('Content-Type') == 'application/json':
            import json
            data = json.loads(request.body)
            if data.get('action') == 'process':
                batch_index = data.get('batch_index', 0)
                file_path = request.session.get('temp_excel_path')
                sheet_name = request.session.get('temp_excel_sheet', 'Listado')
                batch_size = 500 # Procesamos de 500 en 500

                if not file_path or not os.path.exists(file_path):
                    return JsonResponse({'status': 'error', 'message': 'Archivo no encontrado en servidor. Reintente subirlo.'})

                try:
                    # Leer solo el pedazo necesario para no saturar RAM
                    skiprows = (batch_index * batch_size) + 1 # saltamos el header
                    df_batch = pd.read_excel(file_path, engine='openpyxl', sheet_name=sheet_name, usecols="A:F", skiprows=range(1, skiprows), nrows=batch_size, header=0, dtype=str)
                    
                    if df_batch.empty:
                        return JsonResponse({'status': 'progress', 'percent': 100, 'total': 0})

                    df_batch.columns = [str(c).replace(",", "").strip().upper() for c in df_batch.columns]
                    
                    afiliados_to_create = []
                    for _, row in df_batch.iterrows():
                        # 🧼 LIMPIEZA DNI (Preservar ceros a la izquierda y quitar .0)
                        dni_raw = row.get("DNI")
                        if pd.notna(dni_raw):
                            dni = str(dni_raw).strip()
                            if dni.endswith('.0'):
                                dni = dni[:-2]
                        else:
                            dni = None

                        fecha_nac = pd.to_datetime(row.get("FECHA NAC"), dayfirst=True, errors='coerce') if pd.notna(row.get("FECHA NAC")) else None
                        sexo_raw = str(row.get("SEXO")).strip().capitalize() if pd.notna(row.get("SEXO")) else None
                        sexo = sexo_raw[0] if sexo_raw in ["Masculino", "Femenino"] else None
                        
                        afiliados_to_create.append(Afiliado(
                            fecha_nacimiento=fecha_nac,
                            dni=dni,
                            apellido_paterno=str(row.get("APE PATERNO")).strip() if pd.notna(row.get("APE PATERNO")) else None,
                            apellido_materno=str(row.get("APE MATERNO")).strip() if pd.notna(row.get("APE MATERNO")) else None,
                            nombres=str(row.get("NOMBRES")).strip() if pd.notna(row.get("NOMBRES")) else None,
                            sexo=sexo
                        ))
                    
                    Afiliado.objects.bulk_create(afiliados_to_create)
                    
                    # Calcular progreso
                    total_rows = request.session.get('total_excel_rows', 1)
                    processed_so_far = (batch_index + 1) * batch_size
                    percent = min(100, int((processed_so_far / total_rows) * 100))
                    
                    return JsonResponse({
                        'status': 'progress', 
                        'percent': percent,
                        'total_batches': (total_rows // batch_size) + 1
                    })

                except Exception as e:
                    return JsonResponse({'status': 'error', 'message': str(e)})

        # CASO 1: SUBIDA INICIAL -> GENERAR PREVIEW
        elif request.FILES.get("archivo_excel"):
            excel_file = request.FILES["archivo_excel"]
            try:
                # Guardar archivo temporal en /tmp
                import tempfile
                temp_dir = tempfile.gettempdir()
                temp_filename = f"upload_{request.user.id}_{int(timezone.now().timestamp())}.xlsx"
                temp_path = os.path.join(temp_dir, temp_filename)
                
                with open(temp_path, 'wb+') as destination:
                    for chunk in excel_file.chunks():
                        destination.write(chunk)
                
                # Cargar el Excel y buscar la hoja
                xl = pd.ExcelFile(temp_path, engine='openpyxl')
                sheet_name = "Listado" if "Listado" in xl.sheet_names else xl.sheet_names[0]
                
                # Leer conteo total y primeros 10 para preview
                df_full = pd.read_excel(xl, sheet_name=sheet_name, usecols="A:F", header=0, dtype=str)
                total_rows = len(df_full)
                df_preview = df_full.head(10)
                
                # Formatear preview para JSON
                df_preview.columns = [str(c).replace(",", "").strip().upper() for c in df_preview.columns]
                preview_data = []
                for _, row in df_preview.iterrows():
                    # 🧼 LIMPIEZA DNI (Preservar ceros a la izquierda y quitar .0)
                    dni_raw = row.get("DNI")
                    if pd.notna(dni_raw):
                        dni = str(dni_raw).strip()
                        if dni.endswith('.0'):
                            dni = dni[:-2]
                    else:
                        dni = "-"

                    # 🧼 LIMPIEZA FECHA (Quitar 00:00:00)
                    fecha_raw = pd.to_datetime(row.get("FECHA NAC"), dayfirst=True, errors='coerce') if pd.notna(row.get("FECHA NAC")) else None
                    fecha_fmt = fecha_raw.strftime('%d/%m/%Y') if fecha_raw and not pd.isna(fecha_raw) else "-"

                    preview_data.append({
                        'dni': dni,
                        'apellido_paterno': str(row.get("APE PATERNO")).strip() if pd.notna(row.get("APE PATERNO")) else "",
                        'apellido_materno': str(row.get("APE MATERNO")).strip() if pd.notna(row.get("APE MATERNO")) else "",
                        'nombres': str(row.get("NOMBRES")).strip() if pd.notna(row.get("NOMBRES")) else "",
                        'fecha_nac': fecha_fmt,
                        'sexo': str(row.get("SEXO")).strip() if pd.notna(row.get("SEXO")) else "N/A",
                    })

                # Guardar info en sesión
                request.session['temp_excel_path'] = temp_path
                request.session['temp_excel_sheet'] = sheet_name
                request.session['total_excel_rows'] = total_rows
                
                return JsonResponse({
                    'status': 'ok',
                    'total': total_rows,
                    'preview': preview_data,
                    'sheet_used': sheet_name
                })

            except Exception as e:
                return JsonResponse({'status': 'error', 'message': f"Error al leer Excel: {str(e)}"})

    return render(request, "cargar_excel.html")




def procesar_carga_afiliados(request):
    if request.method == "POST":
        try:
            excel_data = request.session.get("excel_data", [])
            total = len(excel_data)
            batch_size = 5000
            processed = 0

            for i in range(0, total, batch_size):
                batch = excel_data[i:i + batch_size]
                afiliados_to_create = []

                for row in batch:
                    sexo_raw = row.get("SEXO", "").strip().capitalize()
                    sexo = sexo_raw[0] if sexo_raw in ["Masculino", "Femenino"] else None

                    afiliado = Afiliado(
                        fecha_nacimiento=row.get("FECHA NAC") or None,
                        dni=row.get("DNI") or None,
                        apellido_paterno=row.get("APE PATERNO") or None,
                        apellido_materno=row.get("APE MATERNO") or None,
                        nombres=row.get("NOMBRES") or None,
                        sexo=sexo
                    )
                    afiliados_to_create.append(afiliado)

                # Bulk insert
                try:
                    Afiliado.objects.bulk_create(afiliados_to_create)
                except IntegrityError as e:
                    # Ignorar duplicados y continuar
                    pass

                processed += len(afiliados_to_create)

            request.session.pop("excel_data", None)
            return JsonResponse({"status": "ok", "message": f"Se han cargado {processed} afiliados."})

        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)})

    return JsonResponse({"status": "error", "message": "Método no permitido"}, status=400)



def login_view(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect("dashboard")
        else:
            messages.error(request, "Usuario o contraseña incorrectos")

    return render(request, "login.html")
def afiliados(request):
    """
    Renderiza el template de Afiliados.
    La tabla se llena vía Ajax desde la vista `afiliados_data`.
    """
    return render(request, "afiliados.html")


def settings(request):

    return render(request, "settings.html")


from django.db.models import Count
from datetime import datetime
@login_required
def dashboard_view(request):
    current_year = datetime.now().year

    # Estadísticas principales
    total_afiliados = Afiliado.objects.count()
    total_atenciones = Atencion.objects.count()
    atenciones_este_anio = Atencion.objects.filter(fecha_atencion__year=current_year).count()

    # Servicios más utilizados este año
    servicios_populares = Servicio.objects.filter(
        atenciones__fecha_atencion__year=current_year
    ).annotate(
        total_atenciones=Count('atenciones')
    ).order_by('-total_atenciones')[:5]

    # Últimas atenciones registradas
    ultimas_atenciones = Atencion.objects.select_related('afiliado', 'servicio').order_by('-fecha_atencion')[:5]

    # Estadísticas de cumplimiento anual
    afiliados_con_servicios = Afiliado.objects.filter(
        atenciones__fecha_atencion__year=current_year
    ).distinct().count()

    context = {
        'total_afiliados': total_afiliados,
        'total_atenciones': total_atenciones,
        'atenciones_este_anio': atenciones_este_anio,
        'servicios_populares': servicios_populares,
        'ultimas_atenciones': ultimas_atenciones,
        'afiliados_con_servicios': afiliados_con_servicios,
        'current_year': current_year,
    }

    return render(request, "dashboard.html", context)


def afiliados_data(request):
    draw = int(request.GET.get('draw', 1))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 50))
    search_value = request.GET.get('search[value]', '')

    anio_actual = datetime.now().year

    queryset = Afiliado.objects.all().order_by('apellido_paterno').prefetch_related('atenciones__servicio')

    if search_value:
        queryset = queryset.filter(
            Q(dni__icontains=search_value) |
            Q(nombres__icontains=search_value) |
            Q(apellido_paterno__icontains=search_value) |
            Q(apellido_materno__icontains=search_value) |
            Q(observacion_enfermeria__icontains=search_value) |
            Q(observacion_odontologia__icontains=search_value) |
            Q(observacion_obstetricia__icontains=search_value)
        )

    total_records = queryset.count()
    page_qs = list(queryset[start:start+length])

    # Precalcular lista de todos los servicios
    servicios_todos = list(Servicio.objects.values_list('nombre', flat=True))

    data = []
    for a in page_qs:
        atenciones = getattr(a, 'atenciones').all()
        servicios_hechos_lista = []
        servicios_hechos_nombres = set()
        for att in atenciones:
            if att.fecha_atencion and getattr(att.fecha_atencion, "year", None) == anio_actual:
                if att.servicio and att.servicio.nombre:
                    servicios_hechos_lista.append({
                        'nombre': att.servicio.nombre,
                        'fecha': att.fecha_atencion.strftime("%d/%m/%Y")
                    })
                    servicios_hechos_nombres.add(att.servicio.nombre)

        # Ordenar por nombre para consistencia
        servicios_hechos_lista.sort(key=lambda x: x['nombre'])
        
        servicios_cumplidos = servicios_hechos_lista if servicios_hechos_lista else "-"
        servicios_pendientes = ', '.join(sorted(set(servicios_todos) - servicios_hechos_nombres)) if servicios_todos else "-"

        data.append({
            'id': a.id,
            'dni': a.dni or "-",
            'apellido_paterno': a.apellido_paterno or "",
            'apellido_materno': a.apellido_materno or "",
            'nombres': a.nombres or "",
            'fecha_nacimiento': a.fecha_nacimiento.strftime("%Y-%m-%d") if a.fecha_nacimiento else "",
            'nombre_completo': f"{a.apellido_paterno or ''} {a.apellido_materno or ''}, {a.nombres or ''}".strip(),
            'sexo': a.get_sexo_display() if a.sexo else "",
            'sexo_val': a.sexo or "",  # para que JS sepa si es "M" o "F"
            'celular': a.celular or "-",
            'servicios_cumplidos': servicios_cumplidos,
            'servicios_pendientes': servicios_pendientes,
            'observacion_enfermeria': a.observacion_enfermeria or "",
            'observacion_odontologia': a.observacion_odontologia or "",
            'observacion_obstetricia': a.observacion_obstetricia or "",
        })

    return JsonResponse({
        'draw': draw,
        'recordsTotal': total_records,
        'recordsFiltered': total_records,
        'data': data
    })


def editar_afiliado(request, afiliado_id):
    afiliado = get_object_or_404(Afiliado, id=afiliado_id)

    if request.method == "POST":
        # Actualizamos a partir del form-urlencoded enviado por $.ajax
        afiliado.dni = request.POST.get("dni") or None
        afiliado.apellido_paterno = request.POST.get("apellido_paterno") or None
        afiliado.apellido_materno = request.POST.get("apellido_materno") or None
        afiliado.nombres = request.POST.get("nombres") or None
        afiliado.sexo = request.POST.get("sexo") or None
        afiliado.celular = request.POST.get("celular") or None

        fecha = request.POST.get("fecha_nacimiento")
        if fecha:
            try:
                afiliado.fecha_nacimiento = datetime.strptime(fecha, "%Y-%m-%d").date()
            except Exception:
                afiliado.fecha_nacimiento = None
        else:
            afiliado.fecha_nacimiento = None

        afiliado.save()
        return JsonResponse({"success": True, "message": "Afiliado actualizado"})

    # GET -> devolver partial HTML con form (ve abajo el partial)
    return render(request, "partials/editar_afiliado_modal.html", {"afiliado": afiliado})


def atenciones_view(request):
    servicios = Servicio.objects.all().order_by("nombre")
    return render(request, "atenciones.html", {"servicios": servicios})

def atenciones_data(request):
    """Devuelve datos para DataTables"""
    atenciones = Atencion.objects.select_related("afiliado", "servicio").all()

    data = []
    for a in atenciones:
        data.append({
            "id": a.id,
            "afiliado": str(a.afiliado),
            "servicio": str(a.servicio),
            "fecha": a.fecha_atencion.strftime("%d/%m/%Y %H:%M"),
            "observaciones": a.observaciones or "-"
        })

    return JsonResponse({"data": data})


# -----------------------------
# BUSCADOR DE AFILIADOS (Select2)
# -----------------------------
def afiliados_search(request):
    q = request.GET.get("q", "").strip()
    if not q:
        return JsonResponse([], safe=False)

    afiliados = Afiliado.objects.filter(
        Q(dni__icontains=q) |
        Q(apellido_paterno__icontains=q) |
        Q(apellido_materno__icontains=q) |
        Q(nombres__icontains=q)
    ).values("id", "dni", "apellido_paterno", "apellido_materno", "nombres")[:20]

    results = []
    for a in afiliados:
        nombre = f"{a['apellido_paterno'] or ''} {a['apellido_materno'] or ''}, {a['nombres'] or ''}".strip()
        results.append({
            "id": a["id"],
            "dni": a["dni"],
            "nombre": nombre
        })
    return JsonResponse(results, safe=False)
@login_required
def get_afiliado_info(request, afiliado_id):
    afiliado = get_object_or_404(Afiliado, id=afiliado_id)
    return JsonResponse({
        'success': True,
        'afiliado': {
            'id': afiliado.id,
            'dni': afiliado.dni or "-",
            'nombre_completo': f"{afiliado.apellido_paterno or ''} {afiliado.apellido_materno or ''}, {afiliado.nombres or ''}".strip(),
            'sexo': afiliado.get_sexo_display() if afiliado.sexo else "-",
            'fecha_nacimiento': afiliado.fecha_nacimiento.strftime("%d/%m/%Y") if afiliado.fecha_nacimiento else "-",
            'celular': afiliado.celular or "-"
        }
    })

# -----------------------------
# LISTA DE ATENCIONES POR AFILIADO
# -----------------------------
def atenciones_list(request, afiliado_id):
    afiliado = get_object_or_404(Afiliado, id=afiliado_id)
    atenciones = Atencion.objects.filter(afiliado=afiliado).select_related("servicio").order_by("-fecha_atencion")

    data = []
    for a in atenciones:
        data.append({
            "servicio": a.servicio.nombre,
            "fecha": a.fecha_atencion.strftime("%d/%m/%Y %H:%M"),
            "observaciones": a.observaciones
        })
    return JsonResponse(data, safe=False)

# -----------------------------
# REGISTRAR NUEVA ATENCIÓN
# -----------------------------
@csrf_exempt
def atenciones(request):
    if request.method == "POST":
        try:
            afiliado_id = request.POST.get("afiliado")
            servicio_id = request.POST.get("servicio")
            observaciones = request.POST.get("observaciones")

            afiliado = get_object_or_404(Afiliado, id=afiliado_id)
            servicio = get_object_or_404(Servicio, id=servicio_id)

            atencion = Atencion.objects.create(
                afiliado=afiliado,
                servicio=servicio,
                observaciones=observaciones
            )

            return JsonResponse({"success": True, "id": atencion.id})

        except Exception as e:
            return JsonResponse({"success": False, "message": str(e)})

    return JsonResponse({"success": False, "message": "Método no permitido"})




# Función para verificar que el usuario es staff/superuser
def is_admin_user(user):
    return user.is_staff or user.is_superuser

@login_required
@user_passes_test(is_admin_user)
def admin_users_search(request):
    """
    View para buscar usuarios (solo para administradores)
    """
    query = request.GET.get('q', '').strip()

    if len(query) < 2:
        return JsonResponse([], safe=False)

    # Buscar SOLO usuarios que NO sean superusuarios
    users = User.objects.filter(
        Q(username__icontains=query) |
        Q(email__icontains=query) |
        Q(first_name__icontains=query) |
        Q(last_name__icontains=query),
        is_superuser=False  # 🔥 clave
    ).order_by('username')[:20]  # Límite de resultados

    users_data = []
    for user in users:
        users_data.append({
            'id': user.id,
            'username': user.username,
            'email': user.email,
            'first_name': user.first_name or '',
            'last_name': user.last_name or '',
            'last_login': user.last_login.strftime('%d/%m/%Y %H:%M') if user.last_login else 'Nunca',
            'is_active': user.is_active,
            'is_staff': user.is_staff,
            'is_superuser': user.is_superuser
        })

    return JsonResponse(users_data, safe=False)


@login_required
@user_passes_test(is_admin_user)
@csrf_exempt
@require_http_methods(["POST"])
def admin_change_password(request):
    """
    View para cambiar contraseña de usuario (solo para administradores)
    """
    try:
        data = json.loads(request.body)
        user_id = data.get('user_id')
        new_password = data.get('new_password')

        # Validaciones básicas
        if not user_id or not new_password:
            return JsonResponse({
                'success': False,
                'message': 'Datos incompletos'
            }, status=400)

        if len(new_password) < 6:
            return JsonResponse({
                'success': False,
                'message': 'La contraseña debe tener al menos 6 caracteres'
            }, status=400)

        # Buscar el usuario
        try:
            user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Usuario no encontrado'
            }, status=404)

        # Cambiar la contraseña usando el método de Django
        user.set_password(new_password)
        user.save()

        # Opcional: Registrar la acción en logs
        print(f"Contraseña cambiada para usuario: {user.username} por admin: {request.user.username}")

        return JsonResponse({
            'success': True,
            'message': f'Contraseña actualizada correctamente para {user.username}'
        })

    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'Error en el formato de datos'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error del servidor: {str(e)}'
        }, status=500)


from django.contrib.auth import update_session_auth_hash

@login_required
@csrf_exempt
@require_http_methods(["POST"])
def change_own_password(request):
    """
    View para que un usuario cambie su propia contraseña
    """
    try:
        # Verificar que el body no esté vacío
        if not request.body:
            return JsonResponse({
                'success': False,
                'message': 'Datos no proporcionados'
            }, status=400)

        data = json.loads(request.body)
        new_password = data.get('new_password')

        if not new_password:
            return JsonResponse({
                'success': False,
                'message': 'La nueva contraseña es requerida'
            }, status=400)

        # Validar nueva contraseña
        if len(new_password) < 6:
            return JsonResponse({
                'success': False,
                'message': 'La nueva contraseña debe tener al menos 6 caracteres'
            }, status=400)

        # Cambiar contraseña
        request.user.set_password(new_password)
        request.user.save()

        # 🔑 MANTENER LA SESIÓN ACTIVA - Esto evita que el usuario se desconecte
        update_session_auth_hash(request, request.user)

        return JsonResponse({
            'success': True,
            'message': 'Contraseña cambiada correctamente'
        })

    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'Error en el formato de datos JSON'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error del servidor: {str(e)}'
        }, status=500)

from django.core.validators import validate_email
from django.core.exceptions import ValidationError
@login_required
@csrf_exempt
@require_http_methods(["POST"])
def change_own_email(request):
    """
    View para que un usuario cambie su propio email
    """
    try:
        # Verificar que el body no esté vacío
        if not request.body:
            return JsonResponse({
                'success': False,
                'message': 'Datos no proporcionados'
            }, status=400)

        data = json.loads(request.body)
        new_email = data.get('new_email')

        if not new_email:
            return JsonResponse({
                'success': False,
                'message': 'El nuevo email es requerido'
            }, status=400)

        # Validar formato de email
        from django.core.validators import validate_email
        from django.core.exceptions import ValidationError
        try:
            validate_email(new_email)
        except ValidationError:
            return JsonResponse({
                'success': False,
                'message': 'El formato del email no es válido'
            }, status=400)

        # Verificar si el email ya existe
        if User.objects.filter(email=new_email).exclude(id=request.user.id).exists():
            return JsonResponse({
                'success': False,
                'message': 'Este email ya está en uso por otro usuario'
            }, status=400)

        # Cambiar email
        request.user.email = new_email
        request.user.save()

        return JsonResponse({
            'success': True,
            'message': 'Email cambiado correctamente'
        })

    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'Error en el formato de datos JSON'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error del servidor: {str(e)}'
        }, status=500)



# Solo superusuarios
def is_superuser(u):
    return u.is_active and u.is_superuser
@login_required
@user_passes_test(is_superuser)
def create_users_from_excel(request):
    """
    Subir .xlsx con columnas en Hoja2:
    username | name | password (DNI)
    """
    stats = {"created": 0, "skipped": 0, "errors": 0, "total_rows": 0}
    logs = []

    if request.method == "POST" and request.FILES.get("excel_file"):
        excel_file = request.FILES["excel_file"]

        try:
            wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)

            if "Hoja2" not in wb.sheetnames:
                messages.error(request, "El archivo no contiene la Hoja2 requerida.")
                return redirect("create_users_excel")

            ws = wb["Hoja2"]
            rows = ws.iter_rows(min_row=2, values_only=True)

            batch = []
            batch_size = 1000
            existing_usernames = set(User.objects.values_list("username", flat=True))

            def sanitize_username(u):
                s = re.sub(r"\s+", " ", (u or "").strip())
                return s[:150]

            for idx, row in enumerate(rows, start=2):  # desde fila 2 por encabezado
                stats["total_rows"] += 1
                try:
                    # 🟩 Ahora esperamos 3 columnas: username, name, password
                    username_raw = row[0] if row and len(row) > 0 else None
                    name_raw = row[1] if row and len(row) > 1 else None
                    password_raw = row[2] if row and len(row) > 2 else None

                    if not username_raw or not password_raw:
                        stats["skipped"] += 1
                        logs.append(f"Fila {idx}: Saltado (faltan username o password)")
                        continue

                    username = sanitize_username(str(username_raw))
                    first_name = str(name_raw or "").strip()

                    if username in existing_usernames:
                        stats["skipped"] += 1
                        logs.append(f"Fila {idx}: Saltado (username ya existe: {username})")
                        continue

                    if isinstance(password_raw, float):
                        password = str(int(password_raw))
                    else:
                        password = str(password_raw).strip()

                    if password == "":
                        stats["skipped"] += 1
                        logs.append(f"Fila {idx}: Saltado (password vacío)")
                        continue

                    email = f"{password}@gmail.com"

                    user = User(
                        username=username,
                        first_name=first_name,
                        email=email,
                        password=make_password(password),
                        is_active=True,
                    )
                    batch.append(user)
                    existing_usernames.add(username)
                    logs.append(f"Fila {idx}: Creado usuario {username} ({first_name}) con email {email}")

                    if len(batch) >= batch_size:
                        User.objects.bulk_create(batch, batch_size=batch_size)
                        stats["created"] += len(batch)
                        batch = []

                except Exception as e_row:
                    stats["errors"] += 1
                    logs.append(f"Fila {idx}: Error → {str(e_row)}")
                    continue

            if batch:
                User.objects.bulk_create(batch, batch_size=batch_size)
                stats["created"] += len(batch)

            return render(request, "create_users_excel.html", {"stats": stats, "logs": logs})

        except Exception as e:
            messages.error(request, f"Error al procesar archivo: {e}")
            return redirect("create_users_excel")

    return render(request, "create_users_excel.html", {"stats": stats, "logs": logs})

@login_required
@user_passes_test(lambda u: u.is_superuser)
def eliminar_todos_afiliados(request):
    if request.method == "POST":
        try:
            total = Afiliado.objects.count()
            Afiliado.objects.all().delete()
            return JsonResponse({'status': 'ok', 'message': f'Se han eliminado {total} afiliados y toda su información relacionada.'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=400)
